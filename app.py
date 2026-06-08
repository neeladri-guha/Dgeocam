from flask import Flask, request, jsonify, render_template
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os, threading
from datetime import datetime

app = Flask(__name__)

EXCEL_FILE = 'reports.xlsx'
HEADERS = [
    '#', 'Agent', 'Report ID', 'Date', 'Site',
    'Category', 'Priority', 'Notes',
    'LAT', 'LNG', 'Accuracy', 'Address',
    'Timestamp', 'Photo Link'
]

lock = threading.Lock()

# ── Create Excel with styled headers if not exists ──────────
def init_excel():
    try:
        if not os.path.exists(EXCEL_FILE):
            raise FileNotFoundError
        load_workbook(EXCEL_FILE)
    except:
        wb = Workbook()
        ws = wb.active
        ws.title = 'GeoCam Reports'
        ws.append(HEADERS)
        for col, _ in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=col)
            cell.font      = Font(bold=True, color='FFFFFF', name='Courier New')
            cell.fill      = PatternFill('solid', fgColor='B71C1C')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        widths = [5, 18, 12, 12, 20, 16, 10, 25, 12, 12, 10, 30, 20, 40]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[1].height = 22
        wb.save(EXCEL_FILE)
# ── Append one row ───────────────────────────────────────────
def append_row(data):
    with lock:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

        row_num = ws.max_row  # next row number (for # column)

        photo_url = data.get('photoURL', '')

        # Parse timestamp
        try:
            ts  = datetime.fromisoformat(data.get('timestamp', ''))
            date_str = ts.strftime('%d %b %Y')
            time_str = ts.strftime('%H:%M:%S')
        except Exception:
            date_str = data.get('date', '')
            time_str = ''

        full_ts = f"{date_str} {time_str}".strip()

        row = [
            row_num,                          # #
            data.get('agent', ''),            # Agent
            data.get('reportId', ''),         # Report ID
            data.get('date', date_str),       # Date
            data.get('site', ''),             # Site
            data.get('category', ''),         # Category
            data.get('priority', ''),         # Priority
            data.get('notes', ''),            # Notes
            data.get('lat', ''),              # LAT
            data.get('lng', ''),              # LNG
            data.get('accuracy', ''),         # Accuracy
            data.get('address', ''),          # Address
            full_ts,                          # Timestamp
            photo_url                         # Photo Link (text first)
        ]

        ws.append(row)

        # Make photo link clickable hyperlink
        last_row = ws.max_row
        link_cell = ws.cell(row=last_row, column=14)
        if photo_url:
            link_cell.hyperlink = photo_url
            link_cell.font = Font(color='2196F3', underline='single', name='Courier New')
            link_cell.value = 'View Photo'

        # Zebra stripe for readability
        if last_row % 2 == 0:
            fill = PatternFill('solid', fgColor='1A0000')
            for col in range(1, len(HEADERS) + 1):
                ws.cell(row=last_row, column=col).fill = fill

        wb.save(EXCEL_FILE)

# ── Routes ───────────────────────────────────────────────────
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/save', methods=['POST'])
def save():
    try:
        data = request.get_json()
        if not data:
            return jsonify({'status': 'error', 'msg': 'No data received'}), 400
        append_row(data)
        return jsonify({'status': 'ok'}), 200
    except Exception as e:
        return jsonify({'status': 'error', 'msg': str(e)}), 500

@app.route('/download')
def download():
    from flask import send_file
    if os.path.exists(EXCEL_FILE):
        return send_file(EXCEL_FILE, as_attachment=True)
    return 'No report file yet.', 404

# ── Start ────────────────────────────────────────────────────
if __name__ == '__main__':
    init_excel()
    app.run(debug=True, host='0.0.0.0', port=5000)
